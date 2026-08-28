import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def scan_music_library(music_root=r"music"):
    """
    扫描 music 文件夹下的所有 IP、歌单、easy/hard 歌曲，返回列表数据
    """
    data = []

    # 检查 music 文件夹是否存在
    if not os.path.exists(music_root):
        print(f"错误：文件夹 '{music_root}' 不存在！")
        return data

    # 遍历一级目录 = IP名称
    for ip_name in os.listdir(music_root):
        ip_path = os.path.join(music_root, ip_name)
        if not os.path.isdir(ip_path):
            continue

        # 遍历二级目录 = 歌单名称
        for playlist_name in os.listdir(ip_path):
            playlist_path = os.path.join(ip_path, playlist_name)
            if not os.path.isdir(playlist_path):
                continue

            # 遍历 easy / hard 文件夹
            for difficulty in ["easy", "hard"]:
                diff_path = os.path.join(playlist_path, difficulty)
                if not os.path.exists(diff_path):
                    continue

                # 遍历所有 mp3 文件
                for filename in os.listdir(diff_path):
                    if filename.lower().endswith(".mp3"):
                        # 去掉后缀名
                        music_name = os.path.splitext(filename)[0]

                        data.append({
                            "IP名称": ip_name,
                            "歌单名称": playlist_name,
                            "难度": difficulty,
                            "歌曲名称": music_name,
                            "文件全名": filename
                        })

    return data

# ===================== Excel 样式常量 =====================
FONT_NAME = "微软雅黑"

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="1F4E79")
_SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color="1F4E79")
_HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
_KEY_FONT = Font(name=FONT_NAME, size=10, bold=True)
_BODY_FONT = Font(name=FONT_NAME, size=10)

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")   # 浅黄：合计行
_EASY_FILL = PatternFill("solid", fgColor="C6EFCE")    # 浅绿：easy
_HARD_FILL = PatternFill("solid", fgColor="FFC7CE")    # 浅红：hard
_EASY_FONT = Font(name=FONT_NAME, size=10, bold=True, color="006100")
_HARD_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")

# 企划（歌单）配色：同一歌单全表同色，超过色板数量时循环取色
_PLAYLIST_FILLS = [
    "DDEBF7",  # 浅蓝
    "FFF2CC",  # 浅黄
    "FCE4D6",  # 浅橙
    "E4DFEC",  # 浅紫
    "DAEEF3",  # 浅青
    "F2DCDB",  # 浅粉
    "E7E6E6",  # 浅灰
    "FBE5D6",  # 浅杏
    "D6DCE4",  # 浅蓝灰
    "EAD1DC",  # 浅玫粉
]

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _display_width(text):
    """估算字符串的显示宽度（中日韩字符按 2 个字符宽计，用于计算列宽）"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(text))


def _safe_sheet_name(name, used_names):
    """生成合法且不重复的工作表名：去非法字符、限长31字符、重名自动加序号"""
    name = "".join(ch for ch in str(name) if ch not in '[]:*?/\\').strip()
    name = name[:31] or "未命名"
    base, n = name, 2
    while name in used_names:
        name = f"{base[:31 - len(f'_{n}')]}_{n}"
        n += 1
    used_names.add(name)
    return name


def _section_header(ws, row, text, span):
    """写入分区标题：合并单元格 + 底纹 + 边框"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    for col in range(1, span + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _SECTION_FILL
        cell.border = _BORDER
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    cell.alignment = _LEFT
    ws.row_dimensions[row].height = 20
    return row + 1


def _set_cell(ws, row, col, value, font=_BODY_FONT, alignment=None, fill=None, border=None):
    """写入一个单元格并应用样式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment or _LEFT
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    return cell


def _playlist_colors(df):
    """为每个歌单分配一个颜色：同一歌单全表同色，颜色不够时循环使用"""
    fills = [PatternFill("solid", fgColor=c) for c in _PLAYLIST_FILLS]
    return {
        name: fills[i % len(fills)]
        for i, name in enumerate(sorted(df["歌单名称"].unique()))
    }


def _write_summary_sheet(wb, df, playlist_colors):
    """写入「总汇总表」：总体统计 + 各IP统计表 + 各歌单明细表"""
    ws = wb.active
    ws.title = "总汇总表"
    ws.sheet_properties.tabColor = "4472C4"

    total_songs = len(df)
    total_ips = df["IP名称"].nunique()
    total_playlists = df.groupby(["IP名称", "歌单名称"]).ngroups

    # 每个 IP 的统计
    ip_summary = (
        df.groupby("IP名称")
        .agg(
            歌单数量=("歌单名称", "nunique"),
            总歌曲数=("歌曲名称", "count"),
            Easy歌曲数=("难度", lambda x: int((x == "easy").sum())),
            Hard歌曲数=("难度", lambda x: int((x == "hard").sum())),
        )
        .reset_index()
    )
    ip_summary["Easy占比"] = (ip_summary["Easy歌曲数"] / ip_summary["总歌曲数"] * 100).round(1)
    # 按总歌曲数从多到少排列，方便快速定位重点 IP
    ip_summary = ip_summary.sort_values("总歌曲数", ascending=False).reset_index(drop=True)

    # 每个 IP + 歌单 + 难度 的明细
    playlist_summary = (
        df.groupby(["IP名称", "歌单名称", "难度"])
        .agg(歌曲数量=("歌曲名称", "count"))
        .reset_index()
    )

    max_cols = 6
    row = 1

    # 标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    _set_cell(ws, row, 1, "音乐库统计总览", font=_TITLE_FONT, alignment=_CENTER)
    ws.row_dimensions[row].height = 30
    row += 2

    # ---- 一、总体统计 ----
    row = _section_header(ws, row, "一、总体统计", max_cols)
    overview = [
        ("总IP数", total_ips),
        ("总歌单数", total_playlists),
        ("总歌曲数", total_songs),
        ("Easy 歌曲数", int((df["难度"] == "easy").sum())),
        ("Hard 歌曲数", int((df["难度"] == "hard").sum())),
    ]
    for name, value in overview:
        _set_cell(ws, row, 1, name, font=_KEY_FONT, border=_BORDER)
        _set_cell(ws, row, 2, value, alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # ---- 二、各IP统计 ----
    row = _section_header(ws, row, "二、各IP统计", max_cols)
    ip_headers = ["IP名称", "歌单数量", "总歌曲数", "Easy歌曲数", "Hard歌曲数", "Easy占比"]
    for col, header in enumerate(ip_headers, start=1):
        _set_cell(ws, row, col, header, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[row].height = 20
    row += 1
    for _, r in ip_summary.iterrows():
        values = [r["IP名称"], r["歌单数量"], r["总歌曲数"], r["Easy歌曲数"], r["Hard歌曲数"], f"{r['Easy占比']}%"]
        for col, value in enumerate(values, start=1):
            _set_cell(ws, row, col, value, alignment=_CENTER if col > 1 else _LEFT, border=_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1
    # 合计行
    easy_total = int(ip_summary["Easy歌曲数"].sum())
    total_vals = [
        "合计",
        int(ip_summary["歌单数量"].sum()),
        int(ip_summary["总歌曲数"].sum()),
        easy_total,
        int(ip_summary["Hard歌曲数"].sum()),
        f"{easy_total / int(ip_summary['总歌曲数'].sum()) * 100:.1f}%",
    ]
    for col, value in enumerate(total_vals, start=1):
        _set_cell(ws, row, col, value, font=_KEY_FONT, fill=_TOTAL_FILL,
                  alignment=_CENTER if col > 1 else _LEFT, border=_BORDER)
    ws.row_dimensions[row].height = 20
    row += 2

    # ---- 三、各歌单详细统计 ----
    row = _section_header(ws, row, "三、各歌单详细统计", max_cols)
    pl_headers = ["IP名称", "歌单名称", "难度", "歌曲数量"]
    for col, header in enumerate(pl_headers, start=1):
        _set_cell(ws, row, col, header, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[row].height = 20
    row += 1
    for _, r in playlist_summary.iterrows():
        pl_fill = playlist_colors[r["歌单名称"]]
        diff_fill = _EASY_FILL if r["难度"] == "easy" else _HARD_FILL
        diff_font = _EASY_FONT if r["难度"] == "easy" else _HARD_FONT
        _set_cell(ws, row, 1, r["IP名称"], fill=pl_fill, border=_BORDER)
        _set_cell(ws, row, 2, r["歌单名称"], font=_KEY_FONT, fill=pl_fill, border=_BORDER)
        _set_cell(ws, row, 3, r["难度"], alignment=_CENTER, font=diff_font, fill=diff_fill, border=_BORDER)
        _set_cell(ws, row, 4, r["歌曲数量"], alignment=_CENTER, fill=pl_fill, border=_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1
    _set_cell(ws, row, 1, "合计", font=_KEY_FONT, fill=_TOTAL_FILL, border=_BORDER)
    _set_cell(ws, row, 2, "", fill=_TOTAL_FILL, border=_BORDER)
    _set_cell(ws, row, 3, "", fill=_TOTAL_FILL, border=_BORDER)
    _set_cell(ws, row, 4, int(playlist_summary["歌曲数量"].sum()),
              font=_KEY_FONT, fill=_TOTAL_FILL, alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[row].height = 20

    # ---- 四、企划（歌单）颜色图例 ----
    row += 2
    row = _section_header(ws, row, "四、企划（歌单）颜色图例", max_cols)
    for col, header in enumerate(["企划名称", "颜色"], start=1):
        _set_cell(ws, row, col, header, font=_HEADER_FONT, fill=_HEADER_FILL,
                  alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[row].height = 20
    row += 1
    for pl_name, pl_fill in playlist_colors.items():
        _set_cell(ws, row, 1, pl_name, font=_KEY_FONT, fill=pl_fill, border=_BORDER)
        _set_cell(ws, row, 2, "", fill=pl_fill, border=_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1

    # 底部：生成时间
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    _set_cell(ws, row, 1, f"生成时间：{datetime.now():%Y-%m-%d %H:%M:%S}",
              font=Font(name=FONT_NAME, size=9, color="808080"), alignment=_LEFT)

    # 列宽与冻结
    ws.column_dimensions["A"].width = 24
    for col in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "A2"


def _write_ip_sheet(wb, df, ip, used_names, playlist_colors):
    """为单个 IP 写入一个 sheet：整行按企划着色 + 难度列按等级着色 + 冻结 + 筛选 + 自适应列宽"""
    ip_df = df[df["IP名称"] == ip].sort_values(
        ["歌单名称", "难度", "歌曲名称"]
    ).reset_index(drop=True)

    ws = wb.create_sheet(title=_safe_sheet_name(ip, used_names))
    ws.sheet_properties.tabColor = "70AD47"

    headers = ["歌单名称", "难度", "歌曲名称", "文件全名"]
    ncols = len(headers)

    # 标题（含 Easy/Hard 统计）
    n_easy = int((ip_df["难度"] == "easy").sum())
    n_hard = int((ip_df["难度"] == "hard").sum())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _set_cell(ws, 1, 1,
              f"IP：{ip} · 歌曲清单（共 {len(ip_df)} 首：Easy {n_easy} ｜ Hard {n_hard}）",
              font=_TITLE_FONT, alignment=_CENTER)
    ws.row_dimensions[1].height = 28

    # 表头
    for col, header in enumerate(headers, start=1):
        _set_cell(ws, 2, col, header, font=_HEADER_FONT, fill=_HEADER_FILL,
                  alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[2].height = 20

    # 数据行：整行按企划（歌单）着色，难度列再按难度等级着色
    r = 3
    for playlist_name, diff, song_name, file_name in ip_df[
        ["歌单名称", "难度", "歌曲名称", "文件全名"]
    ].itertuples(index=False, name=None):
        pl_fill = playlist_colors[playlist_name]
        diff_fill = _EASY_FILL if diff == "easy" else _HARD_FILL
        diff_font = _EASY_FONT if diff == "easy" else _HARD_FONT
        _set_cell(ws, r, 1, playlist_name, font=_KEY_FONT, fill=pl_fill, border=_BORDER)
        _set_cell(ws, r, 2, diff, alignment=_CENTER, font=diff_font, fill=diff_fill, border=_BORDER)
        _set_cell(ws, r, 3, song_name, fill=pl_fill, border=_BORDER)
        _set_cell(ws, r, 4, file_name, fill=pl_fill, border=_BORDER)
        ws.row_dimensions[r].height = 18
        r += 1

    # 自适应列宽
    data_rows = ip_df.to_dict("records")
    for idx, header in enumerate(headers, start=1):
        widths = [_display_width(header)] + [_display_width(row_data[header]) for row_data in data_rows]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max(widths) + 2), 45)

    # 冻结表头 + 自动筛选
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{r - 1}"


def save_to_excel(data, output_file="歌单统计表.xlsx"):
    """保存到 Excel：第一个sheet为总汇总表，每个IP单独一个sheet"""
    if not data:
        print("没有扫描到任何歌曲！")
        return

    df = pd.DataFrame(data)

    wb = Workbook()
    used_names = {"总汇总表"}
    playlist_colors = _playlist_colors(df)
    _write_summary_sheet(wb, df, playlist_colors)
    for ip in sorted(df["IP名称"].unique()):
        _write_ip_sheet(wb, df, ip, used_names, playlist_colors)
    wb.save(output_file)

    print(f"✅ 已生成 Excel：{output_file}")
    print(f"📊 总计歌曲：{len(data)} 首")
    print(f"📁 总IP数：{df['IP名称'].nunique()} 个")

if __name__ == "__main__":
    print("🔍 正在扫描 music 文件夹...")
    music_data = scan_music_library()
    save_to_excel(music_data)
    